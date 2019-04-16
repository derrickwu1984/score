# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

import logging
from openpyxl import Workbook  # 写入Excel表所用
from openpyxl import load_workbook



class ScorePipeline(object):
    title = 'A'
    def __init__(self):
        self.wb = Workbook()
        self.wb = load_workbook('D:\\nm.xlsx')
        self.wb.create_sheet(title=self.title)
        self.ws = self.wb.get_sheet_by_name(self.title)
        # self.ws = self.wb.active
        self.ws.append(['选择批次','选择科类','院校排序方式','选择院校',
                        '表1填报次序', '表1最高分', '表1最低分', '表1录取人数',
                        '表2专业代码', '表2专业名称',
                        '表2填报次序', '表2最高分', '表2最低分', '表2最低分位数', '表2录取人数'])

    def process_item(self, item, spider):
        line = []
        # 表1填报次序
        table1_fill_order_len = len(item['enroll_no'])
        for i in range(len(item['fill_order_table2'])):
            if (i < table1_fill_order_len):
                fill = item['fill_order'][i]
                max = item['max_score'][i]
                min = item['min_score'][i]
                enroll = item['enroll_no'][i]
            else:
                fill = ""
                max = ""
                min = ""
                enroll = ""
            line = [
                    str(item['pcdm'][i]),str(item['kldm'][i]), str(item['pxfs'][i]),
                    str(item['yxdh'][i]),
                    str(fill), str(max), str(min),
                    str(enroll),
                    str(item['pro_code'][i]),str(item['pro_name'][i]),
                    str(item['fill_order_table2'][i]),str(item['max_score_table2'][i]), str(item['min_score_table2'][i]),
                    str(item['min_score_order'][i]),str(item['enroll_no_table2'][i])]
            # self.ws.append(line)  # 将数据以行的形式添加到xlsx中
            # self.wb.save('D:\\nm.xlsx')  # 保存xlsx文件
        return item

